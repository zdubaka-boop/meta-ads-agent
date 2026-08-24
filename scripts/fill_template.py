#!/usr/bin/env python3
"""Turn a structured brief into a filled-in campaign workbook.

Used by the template-populator role: someone dumps copy, ad set names, links
and creative filenames into the chat, Claude turns that into JSON, and this
writes the .xlsx the media buyer will upload.

  python3 scripts/fill_template.py brief.json --out ~/Desktop/CAMPAIGN.xlsx

brief.json:
{
  "account_id": "act_123",                     optional, filled if known
  "campaign":  {"name": "...", "objective": "OUTCOME_TRAFFIC",
                "budget_mode": "CBO", "daily_budget_minor": 2000},
  "defaults":  {"page_id": "...", "link": "...", "cta": "LEARN_MORE",
                "url_tags": "...", "dsa_beneficiary": "..."},
  "adsets": [{"name": "...", "countries": ["LT"], "languages": ["Lithuanian"],
              "age_min": 18, "age_max": 65, "daily_budget_minor": null,
              "optimization_goal": "LINK_CLICKS"}],
  "ads":    [{"adset": "...", "name": "...", "creative": "a.jpg",
              "bodies": ["one","two"], "headlines": ["A"],
              "description": "", "cta": "", "link": ""}]
}

Anything omitted is left blank for a human to fill — nothing is invented.
"""
import argparse, json, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CAMPAIGN_ROWS = {  # Campaign tab is a key/value sheet; these are its row numbers
    "account_id": 2, "campaign_name": 3, "objective": 4, "buying_type": 5,
    "budget_mode": 6, "daily_budget_minor": 7, "lifetime_budget_minor": 8,
    "bid_strategy": 9, "spend_cap_minor": 10, "start_time": 11, "end_time": 12,
    "special_ad_categories": 13, "special_ad_category_country": 14,
    "page_id": 17, "instagram_user_id": 18, "pixel_id": 19, "custom_event_type": 20,
    "link": 21, "display_link": 22, "cta": 23, "url_tags": 24,
    "dsa_beneficiary": 25, "dsa_payor": 26,
}
ADSET_COLS = {  # Ad Sets tab column numbers
    "name": 1, "daily_budget_minor": 2, "lifetime_budget_minor": 3,
    "optimization_goal": 4, "billing_event": 5, "bid_strategy": 6, "bid_amount_minor": 7,
    "countries": 8, "excluded_countries": 9, "cities": 10, "regions": 11,
    "location_types": 12, "languages": 13, "genders": 14, "age_min": 15, "age_max": 16,
    "interests": 17, "excluded_interests": 18, "custom_audiences": 19,
    "excluded_custom_audiences": 20, "advantage_audience": 21, "device_platforms": 22,
    "publisher_platforms": 23, "facebook_positions": 24, "instagram_positions": 25,
    "custom_event_type": 26, "start_time": 27, "end_time": 28, "dsa_beneficiary": 29,
}
AD_COLS = {"adset": 1, "name": 2, "creative": 3, "body": 4, "headline": 5,
           "description": 6, "cta": 7, "link": 8, "display_link": 9,
           "url_tags": 10, "page_id": 11, "instagram_user_id": 12}


def _openpyxl():
    """Import openpyxl, installing it on first use.

    Writing .xlsx needs it, and a media buyer should not have to debug a pip
    error. Every other script in this repo stays standard-library only.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        return load_workbook, Font
    except ImportError:
        pass
    import subprocess
    print("Installing openpyxl (needed once, to write .xlsx files)…")
    for extra in ([], ["--user"], ["--break-system-packages"],
                  ["--user", "--break-system-packages"]):
        try:
            subprocess.run([sys.executable, "-m", "pip", "install", "--quiet",
                            "openpyxl"] + extra, check=True,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            print("  installed.")
            return load_workbook, Font
        except Exception:
            continue
    sys.exit("Could not install openpyxl automatically. Run this once:\n"
             f"  {sys.executable} -m pip install --user openpyxl")


def joined(v):
    """Lists become the sheet's own separators: | for text, comma for codes."""
    if v is None or v == "":
        return None
    if isinstance(v, (list, tuple)):
        v = [str(x).strip() for x in v if str(x).strip()]
        if not v:
            return None
        return " | ".join(v) if any(len(x) > 12 for x in v) else ",".join(v)
    return v


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("brief")
    ap.add_argument("--out", required=True)
    ap.add_argument("--template", default=str(ROOT / "spec" / "CAMPAIGN-TEMPLATE.xlsx"))
    args = ap.parse_args()

    load_workbook, Font = _openpyxl()

    brief = json.loads(Path(args.brief).read_text())
    wb = load_workbook(args.template)
    blue = Font(name="Arial", size=10, color="0000FF")
    put = lambda ws, r, c, v: (ws.cell(row=r, column=c, value=v).__setattr__("font", blue)
                               if v not in (None, "") else None)

    # ---- Campaign tab: clear the examples, write what we were told ----
    c = wb["Campaign"]
    for row in CAMPAIGN_ROWS.values():
        c.cell(row=row, column=2).value = None
    camp, defs = brief.get("campaign", {}), brief.get("defaults", {})
    put(c, CAMPAIGN_ROWS["account_id"], 2, brief.get("account_id"))
    for k in ("campaign_name", "objective", "budget_mode", "daily_budget_minor",
              "lifetime_budget_minor", "special_ad_categories"):
        put(c, CAMPAIGN_ROWS[k], 2, joined(camp.get(k) or camp.get(k.replace("campaign_", ""))))
    for k in ("page_id", "instagram_user_id", "pixel_id", "custom_event_type", "link",
              "display_link", "cta", "url_tags", "dsa_beneficiary", "dsa_payor"):
        put(c, CAMPAIGN_ROWS[k], 2, joined(defs.get(k)))

    # ---- Ad Sets ----
    a = wb["Ad Sets"]
    # Clear past the end of both the template's example rows AND whatever a
    # previous fill left behind, or a smaller re-fill keeps the old tail.
    for row in range(2, max(40, a.max_row + 1, len(brief.get("adsets", [])) + 2)):
        for col in range(1, 30):
            a.cell(row=row, column=col).value = None
    for i, s in enumerate(brief.get("adsets", []), start=2):
        for key, col in ADSET_COLS.items():
            put(a, i, col, joined(s.get(key)))

    # ---- Ads ----
    ads = wb["Ads"]
    for row in range(2, max(400, ads.max_row + 1, len(brief.get("ads", [])) + 2)):
        for col in range(1, 13):
            ads.cell(row=row, column=col).value = None
    for i, ad in enumerate(brief.get("ads", []), start=2):
        put(ads, i, AD_COLS["adset"], ad.get("adset"))
        put(ads, i, AD_COLS["name"], ad.get("name"))
        put(ads, i, AD_COLS["creative"], joined(ad.get("creative")))
        put(ads, i, AD_COLS["body"], joined(ad.get("bodies") or ad.get("body")))
        put(ads, i, AD_COLS["headline"], joined(ad.get("headlines") or ad.get("headline")))
        for k in ("description", "cta", "link", "display_link", "url_tags",
                  "page_id", "instagram_user_id"):
            put(ads, i, AD_COLS[k], joined(ad.get(k)))

    try:
        sys.path.insert(0, str(ROOT / "api" / "_lib"))
        from aitab import add_ai_tab
        add_ai_tab(wb, scope="campaign", context={"account_id": brief.get("account_id")})
    except Exception:
        pass

    out = Path(args.out).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    n_sets = len(brief.get("adsets", []))
    n_ads = len(brief.get("ads", []))
    print(f"Wrote {out}")
    print(f"  campaign : {camp.get('name') or '(blank — needs a name)'}")
    print(f"  {n_sets} ad set(s), {n_ads} ad(s)")
    missing = [k for k in ("page_id", "link", "cta") if not defs.get(k)]
    if missing:
        print(f"  STILL BLANK, must be filled before upload: {', '.join(missing)}")


if __name__ == "__main__":
    main()
