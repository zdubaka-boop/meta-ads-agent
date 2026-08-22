"""Adds an 'AI AGENT — READ FIRST' tab to every workbook the app hands out.

These sheets are filled in by an AI agent, not typed by a person, so the
workbook has to carry its own spec: exact column names, which values are
accepted, what is required, and what must never be invented. Anything the
agent guesses — a budget, a Page, a pixel, a destination URL — is a real
mistake in a real ad account, so those are called out explicitly.
"""
from openpyxl.styles import Font, PatternFill, Alignment

F = "Arial"


def add_ai_tab(wb, *, scope="campaign", context=None):
    """scope: campaign | adsets | ads. context: dict of known values."""
    context = context or {}
    name = "AI AGENT — READ FIRST"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 118
    ws.sheet_view.showGridLines = False

    H = Font(name=F, size=13, bold=True, color="FFFFFF")
    SUB = Font(name=F, size=11, bold=True, color="1F3864")
    N = Font(name=F, size=10)
    M = Font(name="Menlo", size=9.5, color="444444")
    WARN = Font(name=F, size=10, bold=True, color="C00000")
    BAR = PatternFill("solid", fgColor="1F3864")
    r = [1]

    def head(t):
        c = ws.cell(row=r[0], column=1, value=t); c.font = H; c.fill = BAR
        ws.cell(row=r[0], column=2).fill = BAR
        ws.row_dimensions[r[0]].height = 22; r[0] += 2

    def sec(t):
        ws.cell(row=r[0], column=1, value=t).font = SUB; r[0] += 1

    def line(a, b="", font=N):
        ws.cell(row=r[0], column=1, value=a).font = N
        c = ws.cell(row=r[0], column=2, value=b); c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r[0] += 1

    def gap():
        r[0] += 1

    head("AI AGENT — READ THIS BEFORE FILLING ANYTHING IN")

    line("Your job", {
        "campaign": "Fill the Campaign, Ad Sets and Ads tabs to create ONE new campaign.",
        "adsets": "Fill the Ad Sets and Ads tabs ONLY. You are adding ad sets to a campaign "
                  "that already exists. Leave the Campaign tab exactly as it is.",
        "ads": "Fill the Ads tab ONLY. You are adding ads to an ad set that already exists. "
               "Do not touch any other tab.",
    }[scope])
    line("Output", "Save as .xlsx. Do not rename tabs, reorder columns, or add columns.")
    line("Rows in grey", "Already live in the account. Left in as examples. Do not edit them; "
                         "they are skipped on upload. Add your rows BELOW them.")
    gap()

    sec("NEVER INVENT THESE")
    line("", "If a value below is not given to you, STOP and ask. Do not guess, do not copy "
              "from an example, do not use a plausible-looking default. Each one is real "
              "money or a real identity in a live ad account:", WARN)
    line("  daily_budget_minor", "A wrong budget spends real money.")
    line("  page_id / instagram_user_id", "A wrong identity publishes under the wrong brand.")
    line("  pixel_id", "A wrong pixel breaks tracking silently.")
    line("  link", "A wrong destination sends paid traffic to the wrong place.")
    line("  account_id", "A wrong account builds in someone else's ad account.")
    gap()

    sec("FORMAT RULES THAT ARE EASY TO GET WRONG")
    line("Money is in MINOR units", "Whole numbers only. 2000 means 20.00. Never write 20.00 "
                                    "or 20 when you mean twenty. Never use a currency symbol.")
    line("Budget goes in ONE place", "CBO: the budget is on the Campaign tab and every ad set "
                                     "budget must be EMPTY. ABO: the Campaign budget is empty "
                                     "and EVERY ad set needs one. Never both.")
    line("Lists go in one cell", "Comma separated, no spaces needed: GB,IE,DE")
    line("Text variants", "Separate with a pipe: Hook one | Hook two | Hook three. Max 5 "
                          "primary texts and 5 headlines per ad. Commas are NOT separators — "
                          "ad copy is full of commas.")
    line("Several creatives, one row", "Put several names in creative_file separated by | and "
                                       "you get one ad per creative, all sharing the copy.")
    line("Countries", "ISO 3166-1 alpha-2 codes only: GB not UK, GR not EL. See the Countries "
                      "tab. Write 'worldwide' to target everywhere.")
    line("Languages", "The EXACT name from the Languages tab, e.g. 'English (UK)'. Not 'EN', "
                      "not 'English UK'. Leave blank for all languages.")
    line("Dates", "2026-09-01T00:00:00+0300")
    gap()

    sec("TARGETING A LANGUAGE WITH NO COUNTRY")
    line("", "Leave countries EMPTY and fill languages: that targets those speakers worldwide. "
              "You must then put TW,SG in excluded_countries — Meta refuses worldwide delivery "
              "to them without a signed declaration.")
    gap()

    sec("REQUIRED COLUMNS")
    if scope in ("campaign",):
        line("Campaign tab", "account_id, campaign_name, objective, budget_mode (CBO or ABO), "
                             "page_id, link, cta. daily_budget_minor is required for CBO.")
    if scope in ("campaign", "adsets"):
        line("Ad Sets tab", "adset_name, and either countries OR languages. "
                            "optimization_goal defaults to LINK_CLICKS. "
                            "dsa_beneficiary is REQUIRED if any EU country is targeted — it is "
                            "the legal name of the advertiser.")
    line("Ads tab", "adset_name (must match an Ad Sets row EXACTLY, character for character), "
                    "ad_name (unique across the whole sheet), creative_file, body, headline.")
    gap()

    sec("creative_file ACCEPTS ANY OF THESE")
    line("A filename", "e.g. hook_red.jpg — the image must be uploaded alongside the sheet.")
    line("A name already in the account", "The name of an image or video previously uploaded "
                                          "to this ad account. Nothing to attach, no size "
                                          "limit, and this is the ONLY way videos work.")
    line("An image hash", "A 32-character hex string, used as-is.")
    gap()

    sec("VALID VALUES")
    line("objective", "OUTCOME_SALES, OUTCOME_LEADS, OUTCOME_TRAFFIC, OUTCOME_AWARENESS, "
                      "OUTCOME_ENGAGEMENT, OUTCOME_APP_PROMOTION", M)
    line("optimization_goal", "OFFSITE_CONVERSIONS, LANDING_PAGE_VIEWS, LINK_CLICKS, "
                              "IMPRESSIONS, REACH, LEAD_GENERATION, THRUPLAY, VALUE", M)
    line("cta", "LEARN_MORE, SHOP_NOW, SIGN_UP, SUBSCRIBE, GET_OFFER, BUY_NOW, ORDER_NOW, "
                "GET_STARTED, DOWNLOAD, APPLY_NOW, CONTACT_US, NO_BUTTON", M)
    line("genders", "All, Men, Women", M)
    line("publisher_platforms", "facebook, instagram, audience_network, messenger, threads "
                                "(blank = all placements)", M)
    gap()

    if context:
        sec("VALUES ALREADY KNOWN FOR THIS SHEET — USE THESE, DO NOT INVENT ALTERNATIVES")
        for k, v in context.items():
            if v:
                line("  " + k, str(v), M)
        gap()

    sec("BEFORE YOU HAND THE FILE BACK, CHECK")
    line("1", "Every adset_name on the Ads tab matches an Ad Sets row exactly.")
    line("2", "Every ad_name is unique.")
    line("3", "Budgets are whole numbers in minor units, and only on the correct tab.")
    line("4", "Country codes are ISO-2 and language names match the Languages tab exactly.")
    line("5", "You did not invent a budget, page_id, pixel_id, link or account_id.")
    line("6", "The grey example rows are untouched.")
    gap()
    line("", "Everything created from this sheet is PAUSED. A human reviews and launches it. "
              "That is the safety net — do not rely on it to excuse a guess.", WARN)
    return ws
