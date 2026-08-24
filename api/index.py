"""Vercel serverless entrypoint for the Meta Ads Agent console.

Serverless functions are stateless, so there is no in-memory session store.
Instead the cookie carries only an HMAC-signed "you are signed in" flag; the
Meta token stays in the Vercel environment and never reaches the browser.

That means this deployment runs in ACCESS CODE mode only. The own_token mode
(each buyer pastes their own token) stays local-only, because it would require
holding someone else's token in a cookie.

Required env in Vercel:
  META_ACCESS_TOKEN      the Meta token every action runs as
  META_WEB_ACCESS_CODE   the code the team types in
  META_SESSION_SECRET    any long random string; signs the session cookie
"""
import hashlib, hmac, json, os, re, base64, time, traceback, sys
from http.server import BaseHTTPRequestHandler
from pathlib import Path
from urllib.parse import urlparse, parse_qs

sys.path.insert(0, str(Path(__file__).parent / "_lib"))
import meta  # noqa: E402

SECRET = os.getenv("META_SESSION_SECRET", "")
CODE = os.getenv("META_WEB_ACCESS_CODE", "")
TTL = 30 * 24 * 3600      # 30 days; only Sign out ends a session
PUBLIC = Path(__file__).resolve().parent.parent / "public"


BUILD_ID = "0822-1620"

SIG_LEN = 32          # sha256 HMAC, always exactly 32 bytes


def _session_key():
    """Key for sealing the session payload. Derived from the server secret
    only — the browser never has it, so a stolen cookie is inert."""
    import base64 as _b64, hashlib as _h
    return _b64.urlsafe_b64encode(_h.sha256(("sess|" + SECRET).encode()).digest())


def seal(token, who, exp):
    """Encrypt the Meta token into the cookie.

    Reading the token out of blob storage on EVERY request made the whole app
    depend on that storage being up and un-throttled; a single blip logged
    people out mid-task. The token now travels in the cookie, encrypted with
    a server-only key, so storage is touched at login and registration only.
    """
    from cryptography.fernet import Fernet
    payload = json.dumps({"t": token, "w": who, "e": exp})
    return Fernet(_session_key()).encrypt(payload.encode()).decode()


def unseal(blob):
    from cryptography.fernet import Fernet, InvalidToken
    try:
        d = json.loads(Fernet(_session_key()).decrypt(blob.encode()))
    except (InvalidToken, ValueError, TypeError):
        return None
    return d if d.get("e", 0) > time.time() else None


def sign(payload):
    """Sign a session payload.

    The signature is 32 RAW bytes and can contain any byte value, including
    b'.', so it must never be separated from the payload by a delimiter -
    that made roughly one cookie in nine fail its own check. The length is
    fixed, so the split is by position.
    """
    msg = payload.encode() if isinstance(payload, str) else str(payload).encode()
    sig = hmac.new(SECRET.encode(), msg, hashlib.sha256).digest()
    return base64.urlsafe_b64encode(msg + sig).decode().rstrip("=")


def verify(cookie):
    """-> the login code carried by the cookie, or None."""
    if not cookie or not SECRET:
        return None
    try:
        raw = base64.urlsafe_b64decode(cookie + "=" * (-len(cookie) % 4))
        if len(raw) <= SIG_LEN:
            return None
        msg, sig = raw[:-SIG_LEN], raw[-SIG_LEN:]
        good = hmac.new(SECRET.encode(), msg, hashlib.sha256).digest()
        if not hmac.compare_digest(sig, good):
            return None
        exp, code = msg.decode().split("|", 1)
        return code if int(exp) > time.time() else None
    except Exception:
        return None


def unpack_zip(files):
    """A .zip is one drag instead of many. Expand it so a folder can be zipped
    and dropped as a single file, and flatten paths so 'creatives/gold.png'
    still matches a workbook that just says 'gold.png'."""
    import io, zipfile
    out = {}
    for name, blob in files.items():
        if not name.lower().endswith(".zip"):
            out[name] = blob
            continue
        try:
            zf = zipfile.ZipFile(io.BytesIO(blob))
        except zipfile.BadZipFile:
            continue
        for info in zf.infolist():
            if info.is_dir():
                continue
            base = info.filename.rsplit("/", 1)[-1]
            if base.startswith(".") or "__MACOSX" in info.filename:
                continue
            if info.file_size > 12_000_000:
                continue
            out[base] = zf.read(info)
    return out


def _geo_label(targeting):
    """Human label for a targeting block. Worldwide targets carry no
    'countries' key, so this must not assume one."""
    geo = (targeting or {}).get("geo_locations", {}) or {}
    parts = list(geo.get("countries") or [])
    if geo.get("country_groups"):
        parts.append("worldwide")
    if geo.get("cities"):
        parts.append(f"{len(geo['cities'])} city/cities")
    return parts or ["—"]


def insights_by(edge, level, preset):
    """One insights call for a whole level -> {object_id: stats}."""
    out = {}
    key = {"campaign": "campaign_id", "adset": "adset_id", "ad": "ad_id"}[level]
    try:
        # The id field must be requested explicitly, otherwise every row comes
        # back without one and there is nothing to join the stats onto.
        rows = meta.get_all(edge,
            f"{key},spend,impressions,clicks,ctr,cpc,actions,cost_per_action_type",
            limit=200, cap=2000, level=level, date_preset=preset)
    except Exception:
        return out
    for r in rows:
        acts = {a["action_type"]: float(a["value"]) for a in (r.get("actions") or [])}
        cpa_map = {a["action_type"]: float(a["value"])
                   for a in (r.get("cost_per_action_type") or [])}
        pick = lambda m: (m.get("purchase") or m.get("offsite_conversion.fb_pixel_purchase")
                          or m.get("lead") or m.get("link_click"))
        cpa = pick(cpa_map)
        oid = r.get(key)
        if not oid:
            continue
        out[oid] = {"spend": float(r.get("spend") or 0),
                    "impressions": int(r.get("impressions") or 0),
                    "clicks": int(r.get("clicks") or 0),
                    "ctr": round(float(r.get("ctr") or 0), 2),
                    "cpc": round(float(r.get("cpc") or 0), 2),
                    "results": pick(acts) or 0,
                    "cpa": round(cpa, 2) if cpa else None}
    return out


class handler(BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass

    def _send(self, code, payload, ctype="application/json", cookie=None, no_cache=False):
        body = payload if isinstance(payload, bytes) else json.dumps(payload).encode()
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Referrer-Policy", "no-referrer")
        self.send_header("X-Frame-Options", "DENY")
        if no_cache:
            self.send_header("Cache-Control", "no-store, must-revalidate")
            self.send_header("Pragma", "no-cache")
        if not cookie and getattr(self, "_renew", None):
            r = self._renew
            exp = int(time.time()) + TTL
            cookie = (f"ms={seal(r['t'], r.get('w'), exp)}; Path=/; HttpOnly; Secure; "
                      f"SameSite=Lax; Max-Age={TTL}")
        if cookie:
            self.send_header("Set-Cookie", cookie)
        self.end_headers()
        self.wfile.write(body)

    def _fresh(self):
        self._renew = None

    def _sess(self):
        """The signed-in session straight from the cookie. No network."""
        self._fresh()
        m = re.search(r"ms=([A-Za-z0-9_\-=]+)", self.headers.get("Cookie", "") or "")
        return unseal(m.group(1)) if m else None

    def _need_auth(self):
        """Take the Meta token from the sealed cookie. Touches no storage."""
        sess = self._sess()
        if not sess:
            self._send(401, {"error": "not signed in"})
            return False
        os.environ["META_ACCESS_TOKEN"] = sess["t"]
        self._who = sess.get("w")
        self._renew = sess
        return True

    def _body(self):
        n = int(self.headers.get("Content-Length") or 0)
        try:
            return json.loads(self.rfile.read(n)) if n else {}
        except Exception:
            return {}

    # ------------------------------------------------------------------ GET
    def do_GET(self):
        p = urlparse(self.path)
        try:
            if p.path in ("/", "/index.html"):
                # The app shell changes on every deploy, and a cached copy makes
                # people think a fix never shipped. Never cache it.
                return self._send(200, (PUBLIC / "index.html").read_bytes(),
                                  "text/html; charset=utf-8",
                                  no_cache=True)
            if p.path == "/api/sheet":
                # A sheet scoped to where the user is standing, so "which
                # template?" never has to be answered.
                if not self._need_auth():
                    return
                q2 = parse_qs(p.query)
                return self._scoped_sheet((q2.get("scope") or [""])[0],
                                          (q2.get("id") or [""])[0])

            if p.path == "/api/export":
                if not self._need_auth():
                    return
                return self._export_campaign((parse_qs(p.query).get("campaign") or [""])[0])

            if p.path == "/api/template" and (parse_qs(p.query).get("account") or [None])[0]:
                if not self._need_auth():
                    return
                return self._prefilled_template((parse_qs(p.query)["account"])[0])

            if p.path in ("/api/template", "/CAMPAIGN-TEMPLATE.xlsx"):
                # Read BEFORE sending any header: a read failure after the
                # status line is written kills the whole invocation.
                data = None
                for cand in (Path(__file__).parent / "_lib" / "CAMPAIGN-TEMPLATE.xlsx",
                             PUBLIC / "CAMPAIGN-TEMPLATE.xlsx"):
                    try:
                        data = cand.read_bytes(); break
                    except OSError:
                        continue
                if data is None:
                    return self._send(500, {"error": "template file is missing from the deployment"})
                try:
                    import io as _io
                    from openpyxl import load_workbook as _lw
                    from aitab import add_ai_tab as _ai
                    _wb = _lw(_io.BytesIO(data)); _ai(_wb, scope="campaign")
                    _b = _io.BytesIO(); _wb.save(_b); data = _b.getvalue()
                except Exception:
                    pass
                self.send_response(200)
                self.send_header("Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition",
                    'attachment; filename="CAMPAIGN-TEMPLATE.xlsx"')
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return

            if p.path == "/api/session":
                sess = self._sess()
                return self._send(200, {"signed_in": bool(sess), "mode": "code",
                                        "who": (sess or {}).get("w"),
                                        "build": BUILD_ID})
            q = parse_qs(p.query)
            one = lambda k: (q.get(k) or [None])[0]

            if p.path == "/api/accounts":
                if not self._need_auth():
                    return
                S = {1: "ACTIVE", 2: "DISABLED", 3: "UNSETTLED", 7: "PENDING", 9: "GRACE",
                     100: "CLOSED", 101: "ANY_ACTIVE", 201: "ANY_CLOSED"}
                return self._send(200, {"accounts": [{
                    "id": f"act_{a['account_id']}", "name": a.get("name") or a["account_id"],
                    "status": S.get(a.get("account_status"), "?"),
                    "currency": a.get("currency"),
                    "business": (a.get("business") or {}).get("name") or "",
                } for a in meta.ad_accounts()]})

            if p.path == "/api/campaigns":
                if not self._need_auth():
                    return
                camps = meta.get_all(f"{meta.account(one('account'))}/campaigns",
                    "id,name,objective,status,effective_status,daily_budget,"
                    "lifetime_budget,created_time")
                for c in camps:
                    c["budget_mode"] = "CBO" if (c.get("daily_budget") or c.get("lifetime_budget")) else "ABO"
                camps.sort(key=lambda c: c.get("created_time") or "", reverse=True)
                preset = one("preset") or "last_7d"
                stats = insights_by(f"{meta.account(one('account'))}/insights", "campaign", preset)
                for c in camps:
                    c["stats"] = stats.get(c["id"])
                return self._send(200, {"campaigns": camps, "preset": preset,
                                        "has_stats": bool(stats)})

            if p.path == "/api/account-meta":
                if not self._need_auth():
                    return
                i = meta.get(meta.account(one("account")), "name,currency,min_daily_budget")
                return self._send(200, {"currency": i.get("currency"),
                                        "min_daily_budget": i.get("min_daily_budget"),
                                        "name": i.get("name")})

            if p.path == "/api/adsets":
                if not self._need_auth():
                    return
                sets_ = meta.get_all(f"{one('campaign')}/adsets",
                    "id,name,status,effective_status,daily_budget,lifetime_budget,"
                    "optimization_goal,targeting")
                for a in sets_:
                    geo = (a.get("targeting") or {}).get("geo_locations", {}) or {}
                    a["countries"] = geo.get("countries") or []
                    a.pop("targeting", None)
                    try:
                        a["ad_count"] = len(meta.get_all(f"{a['id']}/ads", "id", cap=500))
                    except Exception:
                        a["ad_count"] = None
                preset = one("preset") or "last_7d"
                stats = insights_by(f"{one('campaign')}/insights", "adset", preset)
                for a in sets_:
                    a["stats"] = stats.get(a["id"])
                return self._send(200, {"adsets": sets_, "preset": preset,
                                        "has_stats": bool(stats)})

            if p.path == "/api/countries":
                f = Path(__file__).parent / "_lib" / "reference" / "data" / "countries.json"
                try:
                    return self._send(200, {"countries": json.loads(f.read_text())})
                except OSError:
                    return self._send(200, {"countries": []})

            if p.path == "/api/ad-defaults":
                # Copy from an ad already in this ad set, so "same copy, new
                # images" needs no typing at all.
                if not self._need_auth():
                    return
                out = {}
                # ?ad= reads that exact ad; ?adset= falls back to its first ad.
                FIELDS = "name,creative{object_story_spec,asset_feed_spec,url_tags}"
                one_ad = (one("ad") or "").strip()
                sources = ([meta.get(one_ad, FIELDS)] if one_ad
                           else meta.get_all(f"{one('adset')}/ads", FIELDS, cap=1))
                for ad in sources:
                    cr = ad.get("creative") or {}
                    oss = cr.get("object_story_spec") or {}
                    ld = oss.get("link_data") or oss.get("video_data") or {}
                    feed = cr.get("asset_feed_spec") or {}
                    cta = ld.get("call_to_action") or {}
                    bodies = [b["text"] for b in feed.get("bodies", [])] or \
                             ([ld.get("message")] if ld.get("message") else [])
                    titles = [t["text"] for t in feed.get("titles", [])] or \
                             ([ld.get("name") or ld.get("title")] if
                              (ld.get("name") or ld.get("title")) else [])
                    out = {"body": " | ".join(x for x in bodies if x),
                           "headline": " | ".join(x for x in titles if x),
                           "description": ld.get("description") or "",
                           "cta": cta.get("type") or "LEARN_MORE",
                           "link": ld.get("link") or (cta.get("value") or {}).get("link") or "",
                           "url_tags": cr.get("url_tags") or "",
                           # The Page is the one thing an ad cannot be built
                           # without, and it was never returned before.
                           "page_id": oss.get("page_id") or "",
                           "display": ld.get("caption") or "",
                           "from_ad": ad.get("name")}
                return self._send(200, out)

            if p.path == "/api/ads":
                if not self._need_auth():
                    return
                ads = meta.get_all(f"{one('adset')}/ads",
                    "id,name,status,effective_status,created_time", cap=500)
                # Merge performance so buyers can judge what to switch off
                # without leaving for Ads Manager.
                preset = one("preset") or "last_7d"
                stats = insights_by(f"{one('adset')}/insights", "ad", preset)
                stats_error = None
                for a in ads:
                    a["stats"] = stats.get(a["id"])
                return self._send(200, {"ads": ads, "preset": preset,
                                        "has_stats": bool(stats),
                                        "stats_error": stats_error})

            return self._send(404, {"error": "no such endpoint"})
        except meta.MetaError as e:
            return self._send(400, {"error": str(e)})
        except Exception as e:
            traceback.print_exc()
            return self._send(500, {"error": f"{type(e).__name__}: {e}"})

    def _scoped_sheet(self, scope, oid):
        """A workbook pre-scoped to one campaign or one ad set.

        scope="adsets": for adding ad sets to campaign `oid`. Its existing ad
        sets are listed so the pattern is visible; re-uploading them is
        harmless because duplicates are skipped.
        scope="ads": for adding ads to ad set `oid`. The adset_name column is
        already filled in.
        """
        import io
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        wb = load_workbook(Path(__file__).parent / "_lib" / "CAMPAIGN-TEMPLATE.xlsx")
        blue = Font(name="Arial", size=10, color="0000FF")
        grey = Font(name="Arial", size=10, color="9A9A9A", italic=True)
        note = Font(name="Arial", size=11, bold=True, color="C00000")
        locales = {v: k for k, v in (meta.load_locales() or {}).items()}
        put = lambda ws, r, c, v, f=blue: ws.cell(row=r, column=c, value=v).__setattr__("font", f)

        c = wb["Campaign"]
        for row in range(2, 27):
            c.cell(row=row, column=2).value = None

        a, ads_ws = wb["Ad Sets"], wb["Ads"]
        for row in range(2, 12):
            for col in range(1, 30):
                a.cell(row=row, column=col).value = None
            for col in range(1, 13):
                ads_ws.cell(row=row, column=col).value = None

        if scope == "ads":
            aset = meta.get(oid, "name,campaign{id,name}")
            camp = aset.get("campaign") or {}
            put(c, 2, 2, meta.account((meta.get(camp.get("id"), "account_id") or {})
                                      .get("account_id", "")))
            put(a, 2, 1, aset.get("name"), grey)
            put(a, 3, 1, "^ this ad set already exists — do not change this name", grey)
            existing = meta.get_all(f"{oid}/ads", "name", cap=200)
            r = 2
            for ad in existing[:5]:
                put(ads_ws, r, 1, aset.get("name"), grey)
                put(ads_ws, r, 2, ad.get("name"), grey)
                put(ads_ws, r, 3, "(already live — will be skipped)", grey)
                r += 1
            for _ in range(10):
                put(ads_ws, r, 1, aset.get("name"))
                r += 1
            put(ads_ws, r + 1, 1,
                "Fill the rows above: ad_name, creative_file, body, headline, cta, link.", note)
            put(ads_ws, r + 2, 1,
                "creative_file = a filename you upload, OR the name of an image already "
                "in this ad account.", grey)
            fname = f"ADD-ADS-{(aset.get('name') or 'adset')[:32]}.xlsx"

        else:
            camp = meta.get(oid, "name,account_id,daily_budget,lifetime_budget,objective")
            put(c, 2, 2, meta.account(camp.get("account_id")))
            put(c, 3, 2, camp.get("name"))
            put(c, 4, 2, camp.get("objective"))
            cbo = bool(camp.get("daily_budget") or camp.get("lifetime_budget"))
            put(c, 6, 2, "CBO" if cbo else "ABO")
            put(c, 16, 2, "Leave the rest blank — taken from the live campaign", grey)
            r = 2
            for aset in meta.get_all(f"{oid}/adsets",
                    "name,daily_budget,optimization_goal,billing_event,targeting", cap=20):
                t = aset.get("targeting") or {}
                geo = (t.get("geo_locations") or {})
                put(a, r, 1, aset.get("name"), grey)
                if aset.get("daily_budget"):
                    put(a, r, 2, int(aset["daily_budget"]), grey)
                put(a, r, 4, aset.get("optimization_goal"), grey)
                put(a, r, 5, aset.get("billing_event"), grey)
                put(a, r, 8, ",".join(geo.get("countries") or
                                      (["worldwide"] if geo.get("country_groups") else [])), grey)
                names = [locales.get(x) for x in (t.get("locales") or []) if locales.get(x)]
                if names:
                    put(a, r, 13, ",".join(names), grey)
                put(a, r, 15, t.get("age_min"), grey); put(a, r, 16, t.get("age_max"), grey)
                r += 1
            put(a, r + 1, 1, "^ ad sets above already exist and will be skipped. "
                             "Add YOUR new ad sets in the rows below.", note)
            r += 2
            put(a, r, 1, "")           # first blank row for them
            put(ads_ws, 2, 1, "<- put the new ad set's name here")
            put(ads_ws, 3, 1, "one row per ad, matching an Ad Sets row above", grey)
            fname = f"ADD-ADSETS-{(camp.get('name') or 'campaign')[:32]}.xlsx"

        from aitab import add_ai_tab
        if scope == "ads":
            add_ai_tab(wb, scope="ads", context={
                "target ad set": aset.get("name"),
                "adset_name column": f"must be exactly: {aset.get('name')}",
                "campaign": (camp or {}).get("name")})
        else:
            add_ai_tab(wb, scope="adsets", context={
                "target campaign": camp.get("name"),
                "budget_mode": "CBO" if cbo else "ABO",
                "objective": camp.get("objective"),
                "Campaign tab": "already filled from the live campaign — do not change it"})
        buf = io.BytesIO(); wb.save(buf); data = buf.getvalue()
        safe = "".join(ch if (ch.isalnum() or ch in "-_.") else "-" for ch in fname)
        self.send_response(200)
        self.send_header("Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{safe}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _export_campaign(self, campaign_id):
        """Write an existing campaign back out as a workbook.

        Copying a winner and changing the copy is how buyers actually work, and
        it is Meta's own export/re-import pattern. Creatives come out as image
        hashes, so the re-upload needs no image files at all.
        """
        import io
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        camp = meta.get(campaign_id, "name,objective,status,daily_budget,lifetime_budget,"
                                     "account_id,special_ad_categories")
        acct = meta.account(camp.get("account_id"))
        cbo = bool(camp.get("daily_budget") or camp.get("lifetime_budget"))
        locales = {v: k for k, v in (meta.load_locales() or {}).items()}

        wb = load_workbook(Path(__file__).parent / "_lib" / "CAMPAIGN-TEMPLATE.xlsx")
        blue = Font(name="Arial", size=10, color="0000FF")
        put = lambda ws, r, c, v: ws.cell(row=r, column=c, value=v).__setattr__("font", blue)

        c = wb["Campaign"]
        for row in range(2, 27):
            c.cell(row=row, column=2).value = None
        put(c, 2, 2, acct)
        put(c, 3, 2, f"{camp.get('name')} (copy)")
        put(c, 4, 2, camp.get("objective"))
        put(c, 6, 2, "CBO" if cbo else "ABO")
        if camp.get("daily_budget"):
            put(c, 7, 2, int(camp["daily_budget"]))
        if camp.get("lifetime_budget"):
            put(c, 8, 2, int(camp["lifetime_budget"]))

        a = wb["Ad Sets"]
        for row in range(2, 12):
            for col in range(1, 30):
                a.cell(row=row, column=col).value = None
        ads_ws = wb["Ads"]
        for row in range(2, 12):
            for col in range(1, 13):
                ads_ws.cell(row=row, column=col).value = None

        page_seen, ar, dr = None, 2, 2
        default_cta = default_link = None
        for aset in meta.get_all(f"{campaign_id}/adsets",
                "id,name,daily_budget,lifetime_budget,optimization_goal,billing_event,"
                "targeting,promoted_object,dsa_beneficiary", cap=200):
            t = aset.get("targeting") or {}
            geo = t.get("geo_locations") or {}
            countries = list(geo.get("countries") or [])
            if geo.get("country_groups"):
                countries = ["worldwide"] + countries
            put(a, ar, 1, aset.get("name"))
            if not cbo and aset.get("daily_budget"):
                put(a, ar, 2, int(aset["daily_budget"]))
            put(a, ar, 4, aset.get("optimization_goal"))
            put(a, ar, 5, aset.get("billing_event"))
            put(a, ar, 8, ",".join(countries))
            excl = ((t.get("excluded_geo_locations") or {}).get("countries") or [])
            if excl:
                put(a, ar, 9, ",".join(excl))
            names = [locales.get(x) for x in (t.get("locales") or []) if locales.get(x)]
            if names:
                put(a, ar, 13, ",".join(names))
            g = t.get("genders") or []
            put(a, ar, 14, "Men" if g == [1] else "Women" if g == [2] else "All")
            put(a, ar, 15, t.get("age_min")); put(a, ar, 16, t.get("age_max"))
            if t.get("publisher_platforms"):
                put(a, ar, 23, ",".join(t["publisher_platforms"]))
            if (aset.get("promoted_object") or {}).get("custom_event_type"):
                put(a, ar, 26, aset["promoted_object"]["custom_event_type"])
            if aset.get("dsa_beneficiary"):
                put(a, ar, 29, aset["dsa_beneficiary"])

            for ad in meta.get_all(f"{aset['id']}/ads",
                    "name,creative{object_story_spec,asset_feed_spec}", cap=500):
                cr = ad.get("creative") or {}
                oss = cr.get("object_story_spec") or {}
                ld = oss.get("link_data") or oss.get("video_data") or {}
                feed = cr.get("asset_feed_spec") or {}
                page_seen = page_seen or oss.get("page_id")
                bodies = [b["text"] for b in feed.get("bodies", [])] or \
                         ([ld.get("message")] if ld.get("message") else [])
                titles = [x["text"] for x in feed.get("titles", [])] or \
                         ([ld.get("name") or ld.get("title")] if (ld.get("name") or ld.get("title")) else [])
                cta = ((ld.get("call_to_action") or {}).get("type"))
                link = ld.get("link") or ((ld.get("call_to_action") or {})
                                          .get("value") or {}).get("link")
                put(ads_ws, dr, 1, aset.get("name"))
                put(ads_ws, dr, 2, ad.get("name"))
                put(ads_ws, dr, 3, ld.get("image_hash") or "")   # hash: no file needed
                put(ads_ws, dr, 4, " | ".join(x for x in bodies if x))
                put(ads_ws, dr, 5, " | ".join(x for x in titles if x))
                put(ads_ws, dr, 6, ld.get("description") or "")
                put(ads_ws, dr, 7, cta); put(ads_ws, dr, 8, link)
                default_cta = default_cta or cta
                default_link = default_link or link
                dr += 1
            ar += 1

        if page_seen:
            put(c, 17, 2, page_seen)
        if default_link:
            put(c, 21, 2, default_link)
        if default_cta:
            put(c, 23, 2, default_cta)

        from aitab import add_ai_tab
        add_ai_tab(wb, scope="campaign", context={
            "account_id": acct,
            "this sheet is": "an EXPORT of a live campaign. Uploading it creates a NEW "
                             "campaign; it does not edit the original.",
            "campaign_name": "already suffixed '(copy)' — rename it if you want something else",
            "creative_file": "these are image hashes from this account. Leave them as they "
                             "are and no image files are needed.",
            "page_id": page_seen or "ASK"})
        buf = io.BytesIO(); wb.save(buf); data = buf.getvalue()
        safe = "".join(ch if ch.isalnum() else "-" for ch in (camp.get("name") or "campaign"))[:48]
        self.send_response(200)
        self.send_header("Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{safe}.xlsx"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _prefilled_template(self, acct):
        """Hand back the workbook already filled in for THIS ad account.

        Hunting for account / Page / pixel IDs is the slowest part of filling
        the sheet in, and getting one wrong is the commonest failure. So they
        are written in, and the account's real Pages and pixels are listed on
        their own tab instead of leaving people to guess.
        """
        import io
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        base = Path(__file__).parent / "_lib" / "CAMPAIGN-TEMPLATE.xlsx"
        wb = load_workbook(base)
        blue = Font(name="Arial", size=10, color="0000FF")

        info = meta.get(meta.account(acct), "name,currency,timezone_name")
        assets = meta.account_assets(acct)
        pages = assets.get("pages") if isinstance(assets.get("pages"), list) else []
        pixels = assets.get("pixels") if isinstance(assets.get("pixels"), list) else []
        igs = assets.get("instagram") if isinstance(assets.get("instagram"), list) else []

        c = wb["Campaign"]
        c.cell(row=2, column=2, value=meta.account(acct)).font = blue
        c.cell(row=3, column=2).value = None                      # they name the campaign
        if len(pages) == 1:
            c.cell(row=17, column=2, value=pages[0]["id"]).font = blue
        else:
            c.cell(row=17, column=2).value = None
        c.cell(row=18, column=2).value = igs[0]["id"] if len(igs) == 1 else None
        c.cell(row=19, column=2).value = pixels[0]["id"] if len(pixels) == 1 else None
        for row in (20, 21, 22, 24, 25, 26):                      # clear example copy
            c.cell(row=row, column=2).value = None

        ws = wb.create_sheet("Your account", 1)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 46
        hdr = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="1F3864")
        for i, t in enumerate(["What", "ID (copy this)", "Name"], start=1):
            cell = ws.cell(row=1, column=i, value=t); cell.font = hdr; cell.fill = fill
        r = 2
        ws.cell(row=r, column=1, value="Ad account"); ws.cell(row=r, column=2, value=meta.account(acct))
        ws.cell(row=r, column=3, value=f"{info.get('name')}  ({info.get('currency')}, "
                                       f"{info.get('timezone_name')})"); r += 2
        for label, rows_ in (("Page", pages), ("Instagram", igs), ("Pixel", pixels)):
            for x in rows_:
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=2, value=x["id"])
                ws.cell(row=r, column=3, value=x.get("name") or x.get("username") or "")
                r += 1
            if not rows_:
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=3, value="(none on this account)"); r += 1
            r += 1
        ws.cell(row=r, column=1, value="Budgets are in CENTS")
        ws.cell(row=r, column=3, value="2000 = 20.00 " + (info.get("currency") or "")); r += 1
        ws.cell(row=r, column=1, value="Already filled in")
        ws.cell(row=r, column=3, value="account_id" +
                (", page_id" if len(pages) == 1 else "") +
                (", pixel_id" if len(pixels) == 1 else ""))

        from aitab import add_ai_tab
        add_ai_tab(wb, scope="campaign", context={
            "account_id": meta.account(acct),
            "page_id": pages[0]["id"] if len(pages) == 1 else
                       "; ".join(f'{x["id"]} = {x.get("name","")}' for x in pages) or "ASK",
            "pixel_id": "; ".join(f'{x["id"]} = {x.get("name","")}' for x in pixels) or "none",
            "currency": info.get("currency"),
            "minimum daily budget (minor units)": info.get("min_daily_budget"),
        })
        buf = io.BytesIO(); wb.save(buf); data = buf.getvalue()
        self.send_response(200)
        self.send_header("Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition",
            f'attachment; filename="CAMPAIGN-{meta.account(acct)}.xlsx"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _duplicate_adset(self):
        """Copy an existing ad set inside the same campaign, changing only what
        the user typed.

        Adding an ad set almost always means "the same thing again, different
        country / audience / budget". Meta can copy the ad set and its ads in
        one call, so this is a ten-second job with no spreadsheet. The copy is
        forced PAUSED, and nothing about the original is touched.
        """
        b = self._body()
        src = str(b.get("adset") or "").strip()
        name = (b.get("name") or "").strip()
        if not src or not name:
            return self._send(400, {"error": "ad set and new name are required"})

        info = meta.get(src, "name,campaign_id,daily_budget,targeting,account_id")
        try:
            new_id = meta.copy_object(src, "adset",
                                      campaign_id=info.get("campaign_id"),
                                      deep_copy="true" if b.get("with_ads") else "false",
                                      rename_strategy="NO_RENAME")
        except Exception as e:
            return self._send(400, {"error": f"Meta refused the copy: {e}"})
        if not new_id:
            return self._send(400, {"error": "Meta did not return a new ad set id"})

        changed, warn = ["copied from " + (info.get("name") or src)], []
        params = {"name": name}

        countries = [c.strip().upper() for c in str(b.get("countries") or "").split(",")
                     if c.strip()]
        if countries:
            t = info.get("targeting") or {}
            geo = dict(t.get("geo_locations") or {})
            geo.pop("country_groups", None); geo.pop("cities", None); geo.pop("regions", None)
            geo["countries"] = countries
            t = dict(t); t["geo_locations"] = geo
            params["targeting"] = json.dumps(t)
            changed.append("countries -> " + ", ".join(countries))

        budget = b.get("daily_budget_minor")
        if budget not in (None, "", 0):
            try:
                params["daily_budget"] = str(int(budget))
                changed.append(f"daily budget -> {int(budget)} (minor units)")
            except (TypeError, ValueError):
                warn.append("budget ignored: not a whole number of cents")

        try:
            meta.post(new_id, params, "update_copy")
        except Exception as e:
            return self._send(200, {"ok": True, "id": new_id, "changed": changed,
                                    "warnings": warn + [f"created, but the edits failed: {e}"]})

        after = meta.get(new_id, "name,status,daily_budget,targeting")
        return self._send(200, {"ok": True, "id": new_id, "name": after.get("name"),
                                "status": after.get("status"), "changed": changed,
                                "warnings": warn,
                                "countries": ((after.get("targeting") or {})
                                              .get("geo_locations") or {}).get("countries")})

    def _quick_add(self):
        """Add ads to an existing ad set from a form — no spreadsheet.

        The day-to-day job is "these new images, the copy we already use".
        Making that require a workbook was the wrong shape. One ad per image;
        everything created PAUSED as usual.
        """
        import multipart, builder
        n = int(self.headers.get("Content-Length") or 0)
        if n > 4_300_000:
            return self._send(413, {"error": "Upload too large (~4.5MB cap). Reference "
                                             "creatives already in the account instead."})
        f, files = multipart.parse(self.rfile.read(n), self.headers.get("Content-Type"))
        files = unpack_zip(files)
        adset_id = (f.get("adset") or "").strip()
        acct = (f.get("account") or "").strip()
        if not adset_id or not acct:
            return self._send(400, {"error": "missing ad set or account"})

        page_id = (f.get("page_id") or "").strip()
        if not page_id:
            try:
                pages = meta.get_all(f"{meta.account(acct)}/promote_pages", "id", cap=5)
                if len(pages) == 1:
                    page_id = pages[0]["id"]
            except Exception:
                pass
        link = (f.get("link") or "").strip()
        if not page_id or not link:
            return self._send(400, {"error": "A destination link and a Page are required — "
                                             "neither is ever guessed."})

        names = [k for k in files if not k.lower().endswith((".xlsx", ".xlsm", ".csv", ".zip"))]
        picked = [x.strip() for x in (f.get("existing") or "").split("|") if x.strip()]
        if not names and not picked:
            return self._send(400, {"error": "Add at least one image."})

        prefix = (f.get("prefix") or "ad").strip() or "ad"
        ads = []
        for i, nm in enumerate(sorted(names) + picked, start=1):
            ads.append({"name": f"{prefix} {i:02d} {Path(nm).stem}"[:80], "creative": nm,
                        "body": f.get("body", ""), "headline": f.get("headline", ""),
                        "description": f.get("description", ""),
                        "cta": f.get("cta") or "LEARN_MORE", "link": link,
                        "url_tags": f.get("url_tags", ""), "page_id": page_id})
        for a in ads:
            for src, many in (("body", "bodies"), ("headline", "headlines"),
                              ("description", "descriptions")):
                v = [x.strip() for x in str(a.get(src) or "").split("|") if x.strip()]
                if v:
                    a[src] = v[0]; a[many] = v

        if f.get("mode") != "create":
            return self._send(200, {"ok": True, "preview": True,
                                    "ads": [a["name"] for a in ads], "count": len(ads)})

        # Names already in the ad set are skipped by add_ads_to_adset.
        defaults = {"page_id": page_id, "link": link}
        log = []
        try:
            res = builder.add_ads_to_adset(acct, adset_id, ads, files, defaults, log.append)
        except Exception as e:
            return self._send(400, {"ok": False, "problems": [str(e)], "log": log})
        return self._send(200, {"ok": True, "created": True, **res, "log": log})

    def _add_adsets(self, do_create):
        """Add new ad sets into a campaign that already exists.

        Deliberately separate from /api/create: that builds a whole new
        campaign, this only appends. Nothing existing is recreated, so no
        running ad set loses its learning.
        """
        import multipart, builder
        n = int(self.headers.get("Content-Length") or 0)
        if n > 4_300_000:
            return self._send(413, {"error": "Upload too large (~4.5MB cap). Reference "
                                             "creatives already in the account instead."})
        fields, files = multipart.parse(self.rfile.read(n), self.headers.get("Content-Type"))
        files = unpack_zip(files)
        campaign_id = (fields.get("campaign") or "").strip()
        if not campaign_id:
            return self._send(400, {"error": "missing campaign"})

        book = next(((k, v) for k, v in files.items()
                     if k.lower().endswith((".xlsx", ".xlsm"))), None)
        if not book:
            return self._send(400, {"error": "Drop in the workbook describing the new ad set(s)."})
        creatives = {k: v for k, v in files.items() if k != book[0]}

        # Everything the sheet does not say is taken from the campaign it is
        # being added to, so the Campaign tab can be left entirely blank.
        camp = meta.get(campaign_id, "name,daily_budget,lifetime_budget,account_id")
        inherit = {"account_id": meta.account(camp.get("account_id")),
                   "budget_mode": "CBO" if (camp.get("daily_budget")
                                            or camp.get("lifetime_budget")) else "ABO"}
        try:
            for aset in meta.get_all(f"{campaign_id}/adsets", "id,dsa_beneficiary", cap=5):
                if aset.get("dsa_beneficiary"):
                    inherit["dsa_beneficiary"] = aset["dsa_beneficiary"]
                for ad in meta.get_all(f"{aset['id']}/ads",
                                       "creative{object_story_spec,url_tags}", cap=1):
                    cr = ad.get("creative") or {}
                    oss = cr.get("object_story_spec") or {}
                    ld = oss.get("link_data") or oss.get("video_data") or {}
                    inherit.setdefault("page_id", oss.get("page_id"))
                    cta = (ld.get("call_to_action") or {})
                    inherit.setdefault("cta", cta.get("type"))
                    inherit.setdefault("link", ld.get("link")
                                       or (cta.get("value") or {}).get("link"))
                    if cr.get("url_tags"):
                        inherit.setdefault("url_tags", cr["url_tags"])
                if inherit.get("page_id"):
                    break
        except Exception:
            pass
        inherit = {k: v for k, v in inherit.items() if v}

        spec, problems, resolved = builder.parse_workbook(
            book[1], creatives, mode="append", inherit=inherit)
        if problems:
            return self._send(200, {"ok": False, "problems": problems, "resolved": resolved})

        existing = {a.get("name") for a in
                    meta.get_all(f"{campaign_id}/adsets", "id,name", cap=500)}
        fresh = [a for a in spec["adsets"] if a["name"] not in existing]
        dupes = [a["name"] for a in spec["adsets"] if a["name"] in existing]

        if not do_create:
            return self._send(200, {"ok": True, "preview": True, "resolved": resolved,
                "campaign": camp.get("name"),
                "adsets": [{"name": a["name"],
                            "countries": _geo_label(a["targeting"]),
                            "ads": [x["name"] for x in a["ads"]]} for a in fresh],
                "skipped": dupes,
                "counts": {"adsets": len(fresh),
                           "ads": sum(len(a["ads"]) for a in fresh)}})
        log = []
        try:
            res = builder.add_adsets_to_campaign(
                spec["account_id"], campaign_id, spec, creatives, log.append)
        except builder.PartialBuild as e:
            return self._send(400, {"ok": False, "problems": [str(e)],
                                    "log": log, "partial": e.result})
        except Exception as e:
            return self._send(400, {"ok": False, "problems": [str(e)], "log": log})
        return self._send(200, {"ok": True, "created": True, "result": res, "log": log})

    def _add_ads(self, do_create):
        """Add ads into an ad set that already exists."""
        import multipart, builder
        n = int(self.headers.get("Content-Length") or 0)
        if n > 4_300_000:
            return self._send(413, {"error": "Upload is too large (Vercel caps a request at "
                                             "~4.5MB). Send fewer or smaller images."})
        fields, files = multipart.parse(self.rfile.read(n), self.headers.get("Content-Type"))
        files = unpack_zip(files)
        adset_id = (fields.get("adset") or "").strip()
        acct = (fields.get("account") or "").strip()
        if not adset_id or not acct:
            return self._send(400, {"error": "missing ad set or account"})

        book = next(((k, v) for k, v in files.items()
                     if k.lower().endswith((".xlsx", ".xlsm", ".csv"))), None)
        if not book:
            return self._send(400, {"error": "Drop in a .csv or .xlsx listing the new ads."})
        creatives = {k: v for k, v in files.items() if k != book[0]}

        # Defaults come from the ad set's own campaign context where possible.
        defaults = {}
        try:
            pages = meta.get_all(f"{meta.account(acct)}/promote_pages", "id", cap=5)
            if len(pages) == 1:
                defaults["page_id"] = pages[0]["id"]
        except Exception:
            pass

        ads, problems = builder.parse_ads_only(book[1], book[0], creatives, defaults)
        if problems:
            return self._send(200, {"ok": False, "problems": problems})
        if not do_create:
            return self._send(200, {"ok": True, "preview": True,
                                    "ads": [a["name"] for a in ads],
                                    "count": len(ads)})
        log = []
        try:
            res = builder.add_ads_to_adset(acct, adset_id, ads, creatives, defaults, log.append)
        except Exception as e:
            return self._send(400, {"ok": False, "problems": [str(e)], "log": log})
        return self._send(200, {"ok": True, "created": True, **res, "log": log})

    def _workbook(self, do_create):
        """Parse an uploaded workbook + creatives. Preview, or actually build."""
        import multipart, builder
        n = int(self.headers.get("Content-Length") or 0)
        if n > 4_300_000:
            return self._send(413, {"error":
                "Upload is too large. Vercel caps a serverless request at ~4.5MB. "
                "Send fewer or smaller images, or use the CLI for this batch."})
        fields, files = multipart.parse(self.rfile.read(n), self.headers.get("Content-Type"))
        files = unpack_zip(files)   # a dropped .zip is expanded here

        book = next(((k, v) for k, v in files.items() if k.lower().endswith((".xlsx", ".xlsm"))), None)
        if not book:
            return self._send(400, {"error": "No .xlsx found. Drop the campaign workbook in too."})
        creatives = {k: v for k, v in files.items() if k is not book[0]
                     and not k.lower().endswith((".xlsx", ".xlsm"))}

        spec, problems, resolved = builder.parse_workbook(book[1], creatives)
        if problems:
            return self._send(200, {"ok": False, "problems": problems, "resolved": resolved})

        counts = {"adsets": len(spec["adsets"]),
                  "ads": sum(len(a["ads"]) for a in spec["adsets"])}
        if not do_create:
            return self._send(200, {"ok": True, "preview": True, "resolved": resolved,
                                    "campaign": spec["campaign"], "counts": counts,
                                    "adsets": [{"name": a["name"],
                                                "countries": _geo_label(a["targeting"]),
                                                "ads": [x["name"] for x in a["ads"]]}
                                               for a in spec["adsets"]]})
        log_lines = []
        try:
            res = builder.build(spec, creatives, log_lines.append)
        except builder.PartialBuild as e:
            return self._send(400, {"ok": False, "problems": [str(e)], "log": log_lines,
                                    "partial": e.result})
        except Exception as e:
            return self._send(400, {"ok": False, "problems": [str(e)], "log": log_lines})
        return self._send(200, {"ok": True, "created": True, "result": res,
                                "log": log_lines, "counts": counts})

    # ----------------------------------------------------------------- POST
    def _media_upload(self):
        """One creative in, one Meta reference out.

        The browser sends files one at a time so a 200-file batch is not one
        enormous request that dies at the platform body limit. Nothing about
        an ad is created here — this only puts the asset in the account.
        """
        import multipart
        n = int(self.headers.get("Content-Length") or 0)
        if n > 4_300_000:
            return self._send(413, {"error": "That file is over the ~4.5MB request cap for "
                                             "this deployment. Upload it once with "
                                             "scripts/upload_creatives.py, then reference it "
                                             "by name."})
        fields, files = multipart.parse(self.rfile.read(n), self.headers.get("Content-Type"))
        acct = (fields.get("account") or "").strip()
        if not acct:
            return self._send(400, {"error": "missing account"})
        if not files:
            return self._send(400, {"error": "no file in the request"})
        name, data = next(iter(files.items()))
        if str(name).lower().endswith((".mp4", ".mov", ".m4v", ".webm")):
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=Path(name).suffix, delete=False) as fh:
                fh.write(data)
                tmp = fh.name
            try:
                vid = meta.upload_video(acct, tmp, name=Path(name).stem)
            finally:
                os.unlink(tmp)
            return self._send(200, {"name": name, "video_id": vid})
        h = meta.upload_image_bytes(acct, data, name)
        return self._send(200, {"name": name, "hash": h})

    def _ads_create(self):
        """Create ONE ad from a creative already in the account.

        Deliberately one ad per call: a partial failure then costs one ad, and
        the caller keeps every id it already has. Always PAUSED unless the
        caller explicitly says otherwise — and ACTIVE is refused outright,
        because activation is a separate decision from creation.
        """
        b = self._body()
        acct = (b.get("account") or "").strip()
        adset = (b.get("adset") or "").strip()
        name = (b.get("name") or "").strip()
        cre = (b.get("creative") or "").strip()
        if not (acct and adset and name and cre):
            return self._send(400, {"error": "account, adset, name and creative are all required"})
        if str(b.get("status") or "PAUSED").upper() != "PAUSED":
            return self._send(400, {"error": "This endpoint only creates PAUSED ads. "
                                             "Activate deliberately, in Ads Manager or "
                                             "with scripts/set_status.py --launch."})
        t = b.get("text") or {}
        per = t.get("per") or {}
        pick = lambda k, many: ([per[k]] if per.get(k) else
                                [x for x in (t.get(many) or []) if x])
        bodies = pick("body", "bodies")
        heads = pick("headline", "headlines")
        link = (per.get("link") or t.get("link") or "").strip()
        page = (b.get("page_id") or t.get("page_id") or "").strip()
        if not (bodies and link and page):
            return self._send(400, {"error": "primary text, link and Page are required — "
                                             "none of them is ever guessed."})
        enh = b.get("enhancements")
        dof = None
        if enh is not None:
            on = {k for group in enh.values() for k, v in group.items() if v}
            dof = {"degrees_of_freedom_type": "USER_ENROLLED",
                   "creative_features_spec": {f: {"enroll_status": "OPT_IN" if f in on else "OPT_OUT"}
                                              for f in meta.ENHANCEMENTS_OFF}}
        # enhancements_off=False in both calls: the caller's own opt-in/opt-out
        # map is applied below. Letting the helper stamp its blanket DOF_OFF
        # first would silently overrule what the user actually picked.
        common = dict(link=link, body=bodies[0], headline=(heads or [""])[0],
                      description=per.get("desc") or t.get("desc") or None,
                      cta=(per.get("cta") or t.get("cta") or "LEARN_MORE"),
                      url_tags=(t.get("utm") or None),
                      enhancements_off=False)
        if str(cre).isdigit():
            # video_creative takes no bodies/headlines — variants are image-only.
            creative = meta.video_creative(page, cre, None, **common)
        else:
            creative = meta.image_creative(
                page, cre, bodies=bodies if len(bodies) > 1 else None,
                headlines=heads if len(heads) > 1 else None, **common)
        if dof:
            creative["degrees_of_freedom_spec"] = dof
        ad = meta.create_ad(acct, adset, name, creative)
        return self._send(200, {"id": ad.get("id") if isinstance(ad, dict) else ad,
                                "name": name, "status": "PAUSED"})

    def do_POST(self):
        p = urlparse(self.path)
        try:
            if p.path == "/api/register":
                if not SECRET:
                    return self._send(500, {"error": "META_SESSION_SECRET is not set"})
                import store
                token = (self._body().get("token") or "").strip()
                if not token.startswith("EAA"):
                    return self._send(400, {"error": "That does not look like a Meta token "
                                                     "(they start with EAA)."})
                os.environ["META_ACCESS_TOKEN"] = token
                try:
                    me = meta.whoami()
                except Exception as e:
                    return self._send(401, {"error": f"Meta rejected that token: {e}"})
                d = meta.get("debug_token", None, input_token=token).get("data", {})
                need = {"ads_read", "ads_management", "business_management",
                        "pages_show_list", "pages_read_engagement"}
                missing = sorted(need - set(d.get("scopes", [])))
                if missing:
                    return self._send(403, {"error": "That token is missing permissions: "
                                                     + ", ".join(missing)})
                code = store.new_code()
                store.save(code, token, me.get("name"))
                exp = int(time.time()) + TTL
                return self._send(200, {"ok": True, "code": code, "who": me.get("name"),
                                        "expires_at": d.get("expires_at", 0)},
                    cookie=f"ms={seal(token, me.get('name'), exp)}; Path=/; HttpOnly; Secure; "
                           f"SameSite=Lax; Max-Age={TTL}")

            if p.path == "/api/login":
                if not SECRET:
                    return self._send(500, {"error": "META_SESSION_SECRET is not set"})
                import store
                code = (self._body().get("code") or "").strip().lower()
                rec = store.load(code)
                if not rec:
                    time.sleep(1)
                    return self._send(401, {"error": "No account for that code. "
                                                     "If this is your first time, register instead."})
                exp = int(time.time()) + TTL
                return self._send(200, {"ok": True, "who": rec.get("who")},
                    cookie=f"ms={seal(rec['token'], rec.get('who'), exp)}; Path=/; HttpOnly; "
                           f"Secure; SameSite=Lax; Max-Age={TTL}")

            if p.path == "/api/logout":
                return self._send(200, {"ok": True},
                                  cookie="ms=; Path=/; HttpOnly; Secure; SameSite=Lax; Max-Age=0")

            if p.path == "/api/duplicate-adset":
                if not self._need_auth():
                    return
                return self._duplicate_adset()

            if p.path == "/api/quick-add":
                if not self._need_auth():
                    return
                return self._quick_add()

            if p.path in ("/api/adsets-preview", "/api/adsets-create"):
                if not self._need_auth():
                    return
                return self._add_adsets(p.path.endswith("create"))

            if p.path in ("/api/add-preview", "/api/add-create"):
                if not self._need_auth():
                    return
                return self._add_ads(p.path.endswith("create"))

            if p.path in ("/api/preview", "/api/create"):
                if not self._need_auth():
                    return
                return self._workbook(p.path.endswith("create"))

            if p.path == "/api/status":
                if not self._need_auth():
                    return
                b = self._body()
                ids = [str(i) for i in (b.get("ids") or []) if i]
                want = (b.get("status") or "").upper()
                if want not in ("PAUSED", "ACTIVE"):
                    return self._send(400, {"error": "status must be PAUSED or ACTIVE"})
                if not ids:
                    return self._send(400, {"error": "no ads selected"})
                # Turning ads ON spends real money, so it needs its own explicit
                # confirmation. Turning them OFF never does and does not.
                if want == "ACTIVE" and not b.get("confirm_spend"):
                    return self._send(400, {"error": "Activating ads spends money. "
                                                     "confirm_spend must be true."})
                done, failed = [], []
                for i in ids:
                    try:
                        meta.post(i, {"status": want}, "set_status")
                        done.append(i)
                    except Exception as e:
                        failed.append({"id": i, "error": str(e)})
                return self._send(200, {"ok": not failed, "changed": done,
                                        "failed": failed, "status": want})

            if p.path == "/api/rename":
                if not self._need_auth():
                    return
                b = self._body()
                oid, name = str(b.get("id") or ""), (b.get("name") or "").strip()
                if not oid or not name:
                    return self._send(400, {"error": "id and name are required"})
                meta.post(oid, {"name": name}, "rename")
                return self._send(200, {"ok": True, "id": oid, "name": name})

            if p.path == "/api/audit":
                if not self._need_auth():
                    return
                b = self._body()
                edge = (f"{b['campaign']}/ads" if b.get("campaign")
                        else f"{meta.account(b.get('account'))}/ads")
                ads = meta.get_all(edge, "id,name,status,effective_status,adset{name},"
                                         "campaign{name},creative{id,degrees_of_freedom_spec}",
                                   limit=100, cap=1000)
                rows = []
                for a in ads:
                    cfs = ((a.get("creative") or {}).get("degrees_of_freedom_spec") or {}) \
                            .get("creative_features_spec", {}) or {}
                    rows.append({"id": a["id"], "name": a.get("name", ""),
                                 "status": a.get("effective_status", ""),
                                 "campaign": (a.get("campaign") or {}).get("name", ""),
                                 "adset": (a.get("adset") or {}).get("name", ""),
                                 "on": sorted(k for k, v in cfs.items()
                                              if v.get("enroll_status") == "OPT_IN")})
                return self._send(200, {"total": len(rows),
                                        "flagged": sum(1 for r in rows if r["on"]),
                                        "ads": rows})
            if p.path == "/api/media-upload":
                if not self._need_auth():
                    return
                return self._media_upload()

            if p.path == "/api/ads-create":
                if not self._need_auth():
                    return
                return self._ads_create()

            return self._send(404, {"error": "no such endpoint"})
        except meta.MetaError as e:
            return self._send(400, {"error": str(e)})
        except Exception as e:
            traceback.print_exc()
            return self._send(500, {"error": f"{type(e).__name__}: {e}"})
